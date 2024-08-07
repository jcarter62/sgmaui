import csv
import openpyxl
from openpyxl.utils import get_column_letter
from django.core.paginator import Paginator
from django.urls import reverse
from django.shortcuts import render, redirect
from django.http import HttpResponse
from decouple import config
from home.usersettings import UserSettings
from django.contrib.auth.decorators import login_required
import requests
from usergroups import UserGroups


# Create your views here.
@login_required
def gw_input_params(request):
    context = {}
    search_term = ''
    show_account_chk = 'off'
    from_date = '03/01/2023'
    to_date = ''
    calc_date = ''
    tc_code = ''
    code_code = ''
    cc_exclude = ''

    if UserGroups(request).not_user:
        return redirect(reverse('not-authorized'))

    if request.user.is_authenticated:
        username = request.user.username
    else:
        username = ''

    cookies = request.COOKIES.items()
    for item in cookies:
        if item[0] == 'gc_from_date':
            from_date = item[1]
        if item[0] == 'gc_to_date':
            to_date = item[1]
        if item[0] == 'gc_calc_date':
            calc_date = item[1]
        if item[0] == 'gc_tc_code':
            tc_code = item[1]
        if item[0] == 'gc_code_code':
            code_code = item[1]
        if item[0] == 'gc_cc_exclude':
            cc_exclude = item[1]

    context = {
        "title": "Groundwater Calculation Input",

        "from_date": from_date,
        "to_date": to_date,
        "calc_date": calc_date,
        "tc_code": tc_code,
        "code_code": code_code,
        "cc_exclude": cc_exclude,
        "results": {},
        "postbtnstate": 'disabled',
        "postbtninfo": "Post is disabled until calculation is complete.",
        "enable_buttons": False
    }

    if UserGroups(request).is_admin:
        context['isadmin'] = True

    #
    # handle post request
    if request.method == 'POST':
        # if button value is 'viewresults' then redirect to gwcalc_results
        if request.POST['action'] == 'view':
            return redirect(reverse('gw_view_calc_results'))

        if 'gc_from_date' in request.POST:
            from_date = request.POST['gc_from_date']
        if 'gc_to_date' in request.POST:
            to_date = request.POST['gc_to_date']
        if 'gc_calc_date' in request.POST:
            calc_date = request.POST['gc_calc_date']
        if 'gc_tc_code' in request.POST:
            tc_code = request.POST['gc_tc_code']
        if 'gc_code_code' in request.POST:
            code_code = request.POST['gc_code_code']
        if 'gc_cc_exclude' in request.POST:
            cc_exclude = request.POST['gc_cc_exclude']

        calc_post_param = '0'
        if request.POST['action'] == 'postdata':
            calc_post_param = '1'

        # now we need to calculate.
        url = config('API_URL') + 'gwcalc/calc/' + from_date + '/' + to_date + '/' + \
              calc_date + '/' + tc_code + '/' + code_code + '/' + cc_exclude + '/' + calc_post_param + \
              '/user=' + username

        calc_data = []

        response = requests.get(url)
        if response.status_code == 200:
            rawdata = response.json()
            calc_data = rawdata['data']

        set_postbtn_state(calc_data, context)

    #
    # Now get calculation history to display
    # at bottom of form.
    #
    url = config('API_URL') + 'gwcalc/calc_history'
    response = requests.get(url)
    if response.status_code == 200:
        rawdata = response.json()
        calc_data = rawdata['data']
        context['results'] = calc_data.copy()
    return render(request, 'gwcalc_inputs.html', context=context)


def set_postbtn_state(calc_data, context):
    if len(calc_data) > 0:
        if calc_data[0]['process_state'] == 'calculated' and int(calc_data[0]['errors']) == 0:
            context['postbtnstate'] = ''
            context['postbtninfo'] = 'ok to post.'
        else:
            context['postbtnstate'] = 'disabled'
            context['postbtninfo'] = 'Post is disabled until calculated with no errors.'


# @login_required
def gw_view_calc_results(request):
    context = {}
    search_term = ''
    filter_txt = ''

    if request.user.is_authenticated:
        username = request.user.username
    else:
        username = ''

    cookies = request.COOKIES.items()
    for item in cookies:
        if item[0] == 'search_term':
            search_term = item[1]

    context = {
        "title": "Groundwater Calculation Results",
        "results": {},
        "filter": filter_txt,
    }

    url = config('API_URL') + 'gwcalc/calc_results'

    calc_data = []

    response = requests.get(url)
    if response.status_code == 200:
        rawdata = response.json()
        if search_term > '':
            filter_txt = '- Filtered by: ' + search_term
            for item in rawdata['data']:
                txt = item['description'] + ' ' + item['memo'] + ' ' + item['name_id']
                if search_term in txt:
                    calc_data.append(item)
        else:
            filter_txt = ''
            calc_data = rawdata['data']

    context['results'] = calc_data.copy()
    context['filter'] = filter_txt
    #
    # calculate total amount.
    total_qty = 0
    transaction_count = 0
    account_count = 0
    current_account = 0
    for item in calc_data:
        total_qty = total_qty + float(item['amount'])
        transaction_count = transaction_count + 1
        if item['name_id'] != current_account:
            current_account = item['name_id']
            account_count = account_count + 1


    context['total_qty'] = total_qty
    context['transaction_count'] = transaction_count
    context['account_count'] = account_count

    return render(request, 'view_calc_results.html', context=context)

def gw_export_calc_results(request):
    def quote_wrap(txt):
        return '"' + txt + '"'

    search_term = ''
    filter_txt = ''

    cookies = request.COOKIES.items()
    for item in cookies:
        if item[0] == 'search_term':
            search_term = item[1]

    url = config('API_URL') + 'gwcalc/calc_results'

    calc_data = []

    response = requests.get(url)
    if response.status_code == 200:
        rawdata = response.json()
        if search_term > '':
            filter_txt = '- Filtered by: ' + search_term
            for item in rawdata['data']:
                txt = item['description'] + ' ' + item['memo'] + ' ' + item['name_id']
                if search_term in txt:
                    calc_data.append(item)
        else:
            calc_data = rawdata['data']

    # Export functionality
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="gw_calc_results.csv"'
    writer = csv.writer(response, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC, dialect='excel')
    writer.writerow(['Account', 'Category', 'Description', 'Memo', 'Qty'])

    for item in calc_data:
        writer.writerow([
            int(item['name_id']),
            item['code_id'],
            item['description'],
            item['memo'],
            float(item['amount'])
        ])

    return response

def gw_export_calc_results_xlsx(request):
    def quote_wrap(txt):
        return '"' + txt + '"'

    search_term = ''
    filter_txt = ''

    cookies = request.COOKIES.items()
    for item in cookies:
        if item[0] == 'search_term':
            search_term = item[1]

    url = config('API_URL') + 'gwcalc/calc_results'

    calc_data = []

    response = requests.get(url)
    if response.status_code == 200:
        rawdata = response.json()
        if search_term > '':
            filter_txt = '- Filtered by: ' + search_term
            for item in rawdata['data']:
                txt = item['description'] + ' ' + item['memo'] + ' ' + item['name_id']
                if search_term in txt:
                    calc_data.append(item)
        else:
            calc_data = rawdata['data']

    # Export functionality
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="gw_calc_results.xlsx"'

    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the header row
    header_row = ['Account', 'Category', 'Description', 'Memo', 'Qty']

    for col_num, column_title in enumerate(header_row, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    for row_num, item in enumerate(calc_data, 1):
        for col_num, key in enumerate(['name_id', 'code_id', 'description', 'memo', 'amount'], 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            if key == 'name_id':
                value = int(item[key])
            elif key == 'amount':
                value = float(item[key])
            else:
                value = item[key]
            cell.value = value

    # Save the workbook to the response
    workbook.save(response)

    return response